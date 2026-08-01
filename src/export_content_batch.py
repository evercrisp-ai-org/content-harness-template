#!/usr/bin/env python3
"""
Export Content Batch to Excel

Parses all markdown content files in a batch folder and generates a formatted
Excel (.xlsx) workbook with one sheet per platform (Blog, LinkedIn, Facebook,
YouTube). Designed for easy copy/paste and quick navigation.

Usage:
    python src/export_content_batch.py outputs/drafts/content-batch-2026-02-12/
    python src/export_content_batch.py  # auto-detects most recent batch folder
"""

import os
import re
import sys
import glob
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Brand colors (loaded from brand_config.json)
# ---------------------------------------------------------------------------
DEEP_MUTED_BLUE = "243A4B"
BLUE_SLATE = "5F7483"
ANTIQUE_GOLD = "B08D57"
OFF_WHITE = "F6F7F5"
CHARCOAL = "1E2428"
WARM_GRAY = "9AA3A8"


# ---------------------------------------------------------------------------
# Markdown parser
# ---------------------------------------------------------------------------

def parse_content_file(filepath: str) -> dict:
    """Parse a single markdown content file and extract structured fields."""
    with open(filepath, "r", encoding="utf-8") as f:
        text = f.read()

    result = {
        "source_file": os.path.basename(filepath),
        "title": "",
        "type": "",
        "week": "",
        "theme": "",
        "quarterly_ref": "",
        "strategic_context": "",
        "content": "",
        "visual_assets": "",
        "image_prompts": "",
        "platform": "",
        "week_num": 0,
        "post_num": 0,
        "story_classifications": "",
        "relevance_score": "",
    }

    # --- Title (first H1) ---
    title_match = re.match(r"^#\s+(.+)", text, re.MULTILINE)
    if title_match:
        result["title"] = title_match.group(1).strip()

    # --- Metadata fields ---
    meta_patterns = {
        "type": r"\*\*Type:\*\*\s*(.+)",
        "week": r"\*\*Week:\*\*\s*(.+)",
        "theme": r"\*\*Theme:\*\*\s*(.+)",
        "quarterly_ref": r"\*\*Quarterly plan reference:\*\*\s*(.+)",
        "strategic_context": r"\*\*Strategic context:\*\*\s*(.+)",
        "story_classifications": r"\*\*Story classifications used:\*\*\s*(.+)",
        "relevance_score": r"\*\*Relevance Score:\s*(\w+)\*\*",
    }
    for key, pattern in meta_patterns.items():
        match = re.search(pattern, text)
        if match:
            result[key] = match.group(1).strip()

    # --- Platform detection from filename ---
    fname = os.path.basename(filepath).lower()
    if "-blog-" in fname:
        result["platform"] = "Blog"
    elif "-linkedin-" in fname:
        result["platform"] = "LinkedIn"
    elif "-facebook-" in fname:
        result["platform"] = "Facebook"
    elif "-podcast-" in fname or "-youtube-" in fname:
        result["platform"] = "Podcast"
    elif "-clip-" in fname:
        result["platform"] = "Clips"

    # --- Week number for sorting ---
    week_match = re.search(r"week-(\d+)", fname)
    if week_match:
        result["week_num"] = int(week_match.group(1))

    # --- Post number for social posts and clips ---
    post_match = re.search(r"(?:linkedin|facebook|clip)-(\d+)", fname)
    if post_match:
        result["post_num"] = int(post_match.group(1))

    # --- Visual Assets section ---
    visual_match = re.search(
        r"## Visual Assets?\b[^\n]*\n(.*?)(?=\n---\s*\n(?!##\s*Visual))",
        text,
        re.DOTALL,
    )
    if visual_match:
        raw_visual = visual_match.group(1).strip()
        # Clean up into a readable summary
        result["visual_assets"] = _clean_visual_assets(raw_visual)

    # --- Image prompt(s) (independent of the Visual Assets header spelling) ---
    result["image_prompts"] = _extract_image_prompts(text)

    # --- Content body ---
    # Content is between the last "---" after Visual Assets and "## Quality Checklist"
    result["content"] = _extract_content_body(text)

    return result


def _clean_visual_assets(raw: str) -> str:
    """Condense visual asset block into a readable summary for a spreadsheet cell."""
    sections = []
    current_heading = ""
    current_items = []

    for line in raw.split("\n"):
        line = line.strip()
        if line.startswith("###"):
            if current_heading and current_items:
                sections.append(f"{current_heading}\n" + "\n".join(current_items))
            current_heading = line.lstrip("#").strip()
            current_items = []
        elif line.startswith("- **"):
            # Extract key: value
            kv_match = re.match(r"- \*\*(.+?):\*\*\s*(.*)", line)
            if kv_match:
                current_items.append(f"{kv_match.group(1)}: {kv_match.group(2)}")
        elif line and current_items:
            # Continuation of previous value
            current_items[-1] += " " + line

    if current_heading and current_items:
        sections.append(f"{current_heading}\n" + "\n".join(current_items))

    return "\n\n".join(sections)


def _extract_image_prompts(text: str) -> str:
    """Pull the full 9-point AI image prompt(s) out of a draft file.

    Scans the whole file for '- **AI image prompt ...:** <prompt>' bullets
    (including any wrapped continuation lines) rather than relying on the
    Visual Assets section header, so it works whether a piece uses the plural
    "## Visual Assets" header (blog, podcast) or the singular "## Visual Asset"
    (LinkedIn, Facebook, clips). When a file carries more than one image
    (e.g. a blog hero + in-body infographic), each prompt is labelled with its
    preceding "### ..." slot heading and the prompts are separated by a blank
    line so they read cleanly in a single spreadsheet cell.
    """
    prompt_re = re.compile(r"^-\s*\*\*AI image prompt.*?:\*\*\s*(.*)$", re.IGNORECASE)
    current_heading = ""
    prompts = []  # list of (heading, text)
    capturing = False

    for line in text.split("\n"):
        stripped = line.strip()
        if stripped.startswith("###"):
            current_heading = stripped.lstrip("#").strip()
            capturing = False
            continue
        m = prompt_re.match(stripped)
        if m:
            prompts.append([current_heading, m.group(1).strip()])
            capturing = True
            continue
        if capturing:
            # Stop on a new bullet, heading, divider, or blank line; otherwise
            # treat the line as a wrapped continuation of the current prompt.
            if not stripped or stripped.startswith(("-", "#", "---", "*")):
                capturing = False
            else:
                prompts[-1][1] += " " + stripped

    if not prompts:
        return ""
    if len(prompts) == 1:
        return prompts[0][1]
    return "\n\n".join(
        (f"{heading}: {body}" if heading else body) for heading, body in prompts
    )


def _extract_content_body(text: str) -> str:
    """Extract the main content body from the markdown file.

    File structure (recipe SS12):
      # Title
      ## Post Metadata
      ---
      ## Visual Assets  (or "## Visual Asset", singular, on single-image pieces)
      ...
      ---
      ## Clip Extraction Map   <-- podcast files only, optional
      ---
      ## Content    (clip files use "## Clip Script" instead)
      (content body)
      ---              <-- optional trailing divider
      ## Quality Checklist

    Body extraction anchors directly on the "## Content" / "## Clip Script"
    heading rather than counting dividers, so it's unaffected by optional
    sections (like the podcast's Clip Extraction Map) appearing in between.
    """
    # Strip out the Quality Checklist and everything after it
    parts = re.split(r"\n## Quality Checklist", text, maxsplit=1)
    before_checklist = parts[0]

    # The body always lives under the LAST top-level ("## ", not "### ") section
    # heading before Quality Checklist -- whatever that section happens to be
    # named ("Content", "Clip Script", "Production Brief", etc).
    heading_matches = list(re.finditer(r"(?m)^##\s+.+$", before_checklist))
    if heading_matches:
        content = before_checklist[heading_matches[-1].end():]
    else:
        content = before_checklist

    content = content.strip()

    # Remove trailing "---" that sits right before Quality Checklist
    content = re.sub(r"\n---\s*$", "", content).strip()

    return content


# ---------------------------------------------------------------------------
# Google Drive export (clean, markdown-free text for Docs + Sheets)
# ---------------------------------------------------------------------------
#
# Local .md drafts stay the authoritative, git-tracked source and keep their
# full markdown/metadata/Quality Checklist untouched. These helpers produce a
# separate, cleaned-up representation used only when publishing a finished
# piece to Google Drive, so the Doc a reviewer opens reads as plain prose
# instead of showing literal '#' / '**' characters.

# Weekday each numbered post publishes on (per content-production-batch.mdc).
PUBLISH_DAY_BY_PLATFORM_AND_NUM = {
    ("LinkedIn", 1): 1,   # Tuesday (hook)
    ("LinkedIn", 2): 2,   # Wednesday (story)
    ("LinkedIn", 3): 3,   # Thursday (mechanism)
    ("Facebook", 1): 0,   # Monday
    ("Facebook", 2): 1,   # Tuesday
    ("Facebook", 3): 2,   # Wednesday
    ("Facebook", 4): 3,   # Thursday
    ("Facebook", 5): 4,   # Friday
}

# Blog and Podcast are the week's flagship pieces; both land Monday, alongside
# Facebook post 1. Confirmed with Dave 2026-07-27.
FLAGSHIP_DAY_OFFSET = 0  # Monday

# Clips are repurposed from the podcast and have no fixed slot, so they're
# spread one-per-day Tuesday through Friday, round-robin if there are more
# than 4 in a batch (a 5th clip lands back on Tuesday, taking the Clips-2
# suffix via the existing collision-suffix pattern). Confirmed with Dave 2026-07-27.
CLIPS_DAY_OFFSETS = [1, 2, 3, 4]  # Tue, Wed, Thu, Fri

_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_week_monday(week_str: str):
    """Parse the Monday start date out of a '**Week:**' string like
    'Week 28 (Aug 31 - Sep 6, 2026)'. Returns a date or None if unparseable.
    """
    import datetime

    match = re.search(
        r"\(([A-Za-z]{3,9})\s+(\d{1,2})\s*[–—-]\s*(?:[A-Za-z]{3,9}\s+)?(\d{1,2}),?\s*(\d{4})\)",
        week_str,
    )
    if not match:
        return None
    month_name, day, _end_day, year = match.groups()
    month = _MONTHS.get(month_name[:3].lower())
    if not month:
        return None
    try:
        return datetime.date(int(year), month, int(day))
    except ValueError:
        return None


def _fmt_publish_date(d) -> str:
    """Format a date without a zero-padded day, e.g. 'Tuesday, September 1, 2026'."""
    return f"{d.strftime('%A, %B')} {d.day}, {d.year}"


def clean_markdown_text(text: str) -> str:
    """Strip markdown/formatting syntax, leaving plain readable prose.

    Removes bold/italic markers, stray backslash-escapes some tools introduce
    (\\#, \\-, \\[), and any leading heading markers. Content in brackets like
    [ILLUSTRATIVE] is a compliance tag, not markup, and is left untouched.
    """
    text = re.sub(r"\\([#\-\[\]_*`])", r"\1", text)          # \# -> #, \[ -> [, etc.
    text = re.sub(r"\*\*(.+?)\*\*", r"\1", text)              # **bold**
    text = re.sub(r"(?<!\*)\*([^\*\n]+?)\*(?!\*)", r"\1", text)  # *italic*
    text = re.sub(r"(?m)^#{1,6}\s+", "", text)                # # Heading (requires a space, so #hashtags survive)
    text = re.sub(r"\n{3,}", "\n\n", text).strip()
    return text


def build_doc_export_text(item: Dict) -> str:
    """Build the clean, publish-ready text for a Google Doc upload."""
    lines = [item["title"], ""]
    meta_lines = []
    if item.get("type"):
        meta_lines.append(f"Type: {item['type']}")
    if item.get("week"):
        meta_lines.append(f"Week: {item['week']}")

    day_offset = PUBLISH_DAY_BY_PLATFORM_AND_NUM.get((item["platform"], item["post_num"]))
    monday = _parse_week_monday(item.get("week", "")) if day_offset is not None else None
    if monday is not None:
        import datetime

        publish_date = monday + datetime.timedelta(days=day_offset)
        meta_lines.append(f"Publish Day: {_fmt_publish_date(publish_date)}")
    elif item.get("theme"):
        meta_lines.append(f"Theme: {item['theme']}")

    if meta_lines:
        lines.extend(meta_lines)
        lines.append("")
        lines.append("---")
        lines.append("")
    lines.append(clean_markdown_text(item["content"]))
    return "\n".join(lines).strip() + "\n"


_DAY_ABBR = ["MON", "TUE", "WED", "THU", "FRI", "SAT", "SUN"]

# LinkedIn pieces without a numbered slot (native video, carousel) sit outside
# the three fixed text-post days per the LinkedIn guidelines. Default them to
# Friday of the piece's week so every piece still lands in a day folder.
LINKEDIN_DAY_OFFSET_BY_KIND = {1: 1, 2: 2, 3: 3, "native-video": 4, "carousel": 4}


def _drive_day_folder_name(d) -> str:
    """Render a date as the Drive day-folder name, e.g. '23072026-THU'."""
    return f"{d.day:02d}{d.month:02d}{d.year}-{_DAY_ABBR[d.weekday()]}"


def linkedin_kind(fname: str):
    """Return 1/2/3 for numbered LinkedIn text posts, 'native-video' /
    'carousel' for those pieces, or None if fname isn't a LinkedIn piece."""
    low = os.path.basename(fname).lower()
    if "-linkedin-native-video" in low:
        return "native-video"
    if "-linkedin-carousel" in low:
        return "carousel"
    match = re.search(r"-linkedin-(\d+)", low)
    if match:
        return int(match.group(1))
    return None


def linkedin_publish_date(item: Dict, fname: str):
    """Resolve the publish date for a LinkedIn piece, or None if unresolvable."""
    kind = linkedin_kind(fname)
    if kind is None:
        return None
    offset = LINKEDIN_DAY_OFFSET_BY_KIND.get(kind)
    if offset is None:
        return None
    monday = _parse_week_monday(item.get("week", ""))
    if monday is None:
        return None
    import datetime

    return monday + datetime.timedelta(days=offset)


# Deterministic ordering used to assign the -2/-3 suffix when more than one
# LinkedIn piece lands on the same calendar date (e.g. native-video and
# carousel both default to Friday of the same week).
_KIND_ORDER = {1: 0, 2: 1, 3: 2, "native-video": 3, "carousel": 4}


def resolve_publish_date_and_order(item: Dict, fname: str):
    """Resolve (publish_date, order_key) for ANY platform's piece, or (None, None)
    if unresolvable. order_key is used only to break ties when more than one
    piece of the same platform lands on the same date (assigns the -2/-3
    folder suffix in a stable, repeatable order)."""
    import datetime

    platform = item.get("platform", "")
    monday = _parse_week_monday(item.get("week", ""))
    if monday is None:
        return None, None

    if platform == "LinkedIn":
        kind = linkedin_kind(fname)
        offset = LINKEDIN_DAY_OFFSET_BY_KIND.get(kind)
        order_key = _KIND_ORDER.get(kind, 99)
    elif platform == "Facebook":
        offset = PUBLISH_DAY_BY_PLATFORM_AND_NUM.get(("Facebook", item.get("post_num")))
        order_key = item.get("post_num", 99)
    elif platform in ("Blog", "Podcast"):
        offset = FLAGSHIP_DAY_OFFSET
        order_key = 0
    elif platform == "Clips":
        post_num = item.get("post_num") or 0
        if post_num < 1:
            return None, None
        offset = CLIPS_DAY_OFFSETS[(post_num - 1) % len(CLIPS_DAY_OFFSETS)]
        order_key = post_num
    else:
        return None, None

    if offset is None:
        return None, None
    return monday + datetime.timedelta(days=offset), order_key


def _drive_week_folder_name(item: Dict) -> str:
    """Render the Drive week-folder name, reusing the batch's own 'Week N
    (date range)' label verbatim so it matches what's already in every
    piece's metadata rather than introducing a second week-numbering scheme."""
    return item.get("week", "").strip() or "Unscheduled week"


def build_drive_publish_plan(batch_folder: str) -> List[Dict]:
    """Build the Drive publish plan for every piece in a batch folder, across
    all 5 channels (Blog, Podcast, Clips, LinkedIn, Facebook).

    Folder structure is week-first: <Week N (date range)>/<N-DDMMYYYY-DAY>/<Platform[-N]>/,
    with the cleaned copy Doc and the asset file(s) sitting flat together in
    that one folder. A new folder (the -N suffix on the platform name) is
    only created when a later, distinct piece of the SAME platform lands on
    the same date — copy and its own asset never split into separate
    subfolders. The day folder carries a 1-5 (Mon-Fri) numeric prefix so Drive
    sorts the week's days in the right order (DDMMYYYY-DAY alone sorts wrong,
    e.g. Monday's 31082026 sorts after Tuesday's 01092026).
    """
    entries = []
    md_files = sorted(glob.glob(os.path.join(batch_folder, "week-*.md")))
    images_dir = os.path.join(batch_folder, "images")
    for filepath in md_files:
        fname = os.path.basename(filepath)
        item = parse_content_file(filepath)
        if not item.get("platform"):
            continue  # not a recognized content piece (e.g. a stray file)
        publish_date, order_key = resolve_publish_date_and_order(item, fname)
        if publish_date is None:
            continue
        stem = os.path.splitext(fname)[0]
        assets = sorted(glob.glob(os.path.join(images_dir, stem + "*.png")))
        entries.append({
            "source_file": fname,
            "title": item["title"],
            "platform": item["platform"],
            "publish_date": publish_date,
            "order_key": order_key,
            "assets": assets,
            "week_folder": _drive_week_folder_name(item),
        })

    # Group by (date, platform), then assign a stable platform-folder suffix
    # within each group so colliding pieces (same date, same platform) get
    # distinct folders. Different platforms on the same date never collide
    # with each other — each gets its own platform folder regardless.
    by_date_platform: Dict = {}
    for entry in entries:
        by_date_platform.setdefault((entry["publish_date"], entry["platform"]), []).append(entry)

    plan = []
    for (publish_date, platform), group_entries in by_date_platform.items():
        group_entries.sort(key=lambda e: e["order_key"])
        for i, entry in enumerate(group_entries):
            platform_folder = platform if i == 0 else f"{platform}-{i + 1}"
            plan.append({
                "source_file": entry["source_file"],
                "title": entry["title"],
                "platform": platform,
                "publish_date": publish_date.isoformat(),
                "drive_week_folder": entry["week_folder"],
                "drive_date_folder": f"{publish_date.weekday() + 1}-{_drive_day_folder_name(publish_date)}",
                "drive_platform_folder": platform_folder,
                "assets": entry["assets"],
            })
    plan.sort(key=lambda p: (p["publish_date"], p["drive_platform_folder"]))
    return plan


# ---------------------------------------------------------------------------
# Summary sheet (feeds the n8n image-generation workflow's Drive trigger)
# ---------------------------------------------------------------------------
#
# generate-batch's Step 7 uploads a native Google Sheet of this content into
# the dedicated "Summary Sheets" Drive folder once every piece's Doc exists.
# The n8n workflow (automation/n8n-image-generation-workflow.json) watches
# that folder, reads this sheet's File + Image Prompt columns, generates each
# image, and drops it beside the copy Doc the File column points to.

def build_drive_summary_rows(content_items: List[Dict], url_map: Dict[str, str]) -> List[Dict]:
    """Build summary-sheet rows for the n8n image-generation workflow, one row
    per INDIVIDUAL image prompt (not per file). A piece with multiple images
    (e.g. a blog's hero + in-body infographics) gets one row per prompt, all
    sharing the same File URL, so the n8n workflow (one image per row) still
    generates every image for that piece and drops them all in the same
    folder as its Doc. A piece with no image prompt is skipped entirely,
    since there is nothing for the workflow to generate."""
    rows = []
    for item in content_items:
        content_clean = clean_markdown_text(item.get("content", ""))
        preview = content_clean.replace("\n", " ").strip()
        if len(preview) > 300:
            preview = preview[:300].rstrip() + "..."

        publish_date, _ = resolve_publish_date_and_order(item, item["source_file"])
        publish_day = _fmt_publish_date(publish_date) if publish_date else ""

        raw_prompts = item.get("image_prompts", "")
        prompts = [p.strip() for p in raw_prompts.split("\n\n") if p.strip()]
        if not prompts:
            continue

        for prompt in prompts:
            rows.append({
                "File": url_map.get(item["source_file"], ""),
                "Platform": item["platform"],
                "Title": item["title"],
                "Type": item["type"],
                "Week": item["week"],
                "Theme": item["theme"],
                "Publish Day": publish_day,
                "Word Count": len(content_clean.split()),
                "Relevance Score": item.get("relevance_score", ""),
                "Story Classifications": item.get("story_classifications", ""),
                "Strategic Context": item.get("strategic_context", ""),
                "Content Preview": preview,
                "Image Prompt": prompt,
            })
    return rows


def write_drive_summary_csv(rows: List[Dict]) -> str:
    """Render summary rows as CSV text, ready to upload to Drive (auto-converts
    to a Google Sheet with plain-text URLs, no HYPERLINK formulas)."""
    import csv
    import io

    columns = [
        "File", "Platform", "Title", "Type", "Week", "Theme", "Publish Day",
        "Word Count", "Relevance Score", "Story Classifications",
        "Strategic Context", "Content Preview", "Image Prompt",
    ]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=columns)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

def create_workbook(content_items: List[Dict]) -> Workbook:
    """Create a formatted Excel workbook from parsed content items."""
    wb = Workbook()

    # Platform order
    platforms = ["Blog", "Podcast", "Clips", "LinkedIn", "Facebook"]

    # Column definitions
    columns = [
        ("Week", 22),
        ("Title", 40),
        ("Theme", 28),
        ("Type", 22),
        ("Strategic Context", 55),
        ("Content", 85),
        ("Visual Asset Brief", 65),
        ("Image Prompt", 90),
        ("Source File", 38),
    ]

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color=OFF_WHITE)
    header_fill = PatternFill(start_color=DEEP_MUTED_BLUE, end_color=DEEP_MUTED_BLUE, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    subheader_font = Font(name="Calibri", bold=True, size=11, color=CHARCOAL)
    subheader_fill = PatternFill(start_color=WARM_GRAY, end_color=WARM_GRAY, fill_type="solid")

    body_font = Font(name="Calibri", size=11, color=CHARCOAL)
    body_alignment = Alignment(vertical="top", wrap_text=True)

    title_font = Font(name="Calibri", bold=True, size=11, color=DEEP_MUTED_BLUE)

    gold_font = Font(name="Calibri", bold=True, size=11, color=ANTIQUE_GOLD)

    thin_border = Border(
        bottom=Side(style="thin", color=WARM_GRAY),
        top=Side(style="thin", color=WARM_GRAY),
        left=Side(style="thin", color=WARM_GRAY),
        right=Side(style="thin", color=WARM_GRAY),
    )

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for platform in platforms:
        # Filter items for this platform
        items = [i for i in content_items if i["platform"] == platform]
        if not items:
            continue

        # Sort by week number, then post number
        items.sort(key=lambda x: (x["week_num"], x["post_num"]))

        ws = wb.create_sheet(title=platform)

        # --- Header row ---
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        ws.row_dimensions[1].height = 30

        # --- Data rows ---
        current_week = None
        row_idx = 2

        for item in items:
            # Insert week separator row if week changed
            if item["week_num"] != current_week:
                if current_week is not None:
                    # Add a blank spacer row between weeks
                    row_idx += 1

                # Week header row
                week_label = f"WEEK {item['week_num']}"
                if item["week"]:
                    week_label = item["week"].upper()

                cell = ws.cell(row=row_idx, column=1, value=week_label)
                cell.font = subheader_font
                cell.fill = subheader_fill
                for c in range(1, len(columns) + 1):
                    ws.cell(row=row_idx, column=c).fill = subheader_fill
                    ws.cell(row=row_idx, column=c).border = thin_border
                ws.row_dimensions[row_idx].height = 24
                current_week = item["week_num"]
                row_idx += 1

            # Content row
            row_data = [
                item["week"],
                item["title"],
                item["theme"],
                item["type"],
                item["strategic_context"],
                item["content"],
                item["visual_assets"],
                item["image_prompts"],
                item["source_file"],
            ]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.alignment = body_alignment
                cell.border = thin_border

            # Make title column bold with brand color
            ws.cell(row=row_idx, column=2).font = title_font

            # Make week column use gold accent
            ws.cell(row=row_idx, column=1).font = gold_font

            # Set row height based on content length
            content_len = len(item.get("content", ""))
            if content_len > 2000:
                ws.row_dimensions[row_idx].height = 200
            elif content_len > 500:
                ws.row_dimensions[row_idx].height = 120
            else:
                ws.row_dimensions[row_idx].height = 80

            row_idx += 1

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Auto-filter on header row
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{row_idx - 1}"

    return wb


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def find_latest_batch_folder(base_dir: str) -> Optional[str]:
    """Find the most recent content-batch-* folder in the drafts directory."""
    pattern = os.path.join(base_dir, "outputs", "drafts", "content-batch-*")
    folders = sorted(glob.glob(pattern), reverse=True)
    return folders[0] if folders else None


def extract_batch_date(folder_path: str) -> str:
    """Extract the date from a batch folder name and return as YYYYMMDD."""
    folder_name = os.path.basename(folder_path.rstrip("/"))
    date_match = re.search(r"content-batch-(\d{4})-(\d{2})-(\d{2})", folder_name)
    if date_match:
        return f"{date_match.group(1)}{date_match.group(2)}{date_match.group(3)}"
    return "unknown"


def main():
    # --clean-for-drive <file.md>: print clean, markdown-free Doc text to stdout.
    if len(sys.argv) > 1 and sys.argv[1] == "--clean-for-drive":
        if len(sys.argv) < 3:
            print("Usage: export_content_batch.py --clean-for-drive <file.md>")
            sys.exit(1)
        item = parse_content_file(sys.argv[2])
        print(build_doc_export_text(item), end="")
        return

    # --drive-plan <batch_folder>: print JSON describing where each piece,
    # across all 5 channels, belongs in the
    # <Week N>/<D-DDMMYYYY-DAY>/<Platform[-N]>/ Drive tree.
    if len(sys.argv) > 1 and sys.argv[1] == "--drive-plan":
        if len(sys.argv) < 3:
            print("Usage: export_content_batch.py --drive-plan <batch_folder>")
            sys.exit(1)
        import json

        batch_folder = os.path.abspath(sys.argv[2])
        plan = build_drive_publish_plan(batch_folder)
        print(json.dumps(plan, indent=2))
        return

    # --drive-summary-csv <batch_folder> <url_map.json>: print CSV summary rows
    # to stdout, where url_map.json is {"source_file.md": "https://docs.google.com/..."}
    # mapping each piece to the Doc URL Step 7 just uploaded it as. Upload the
    # output as a native Sheet (contentMimeType: text/csv) into the Summary
    # Sheets Drive folder to trigger the n8n image-generation workflow.
    if len(sys.argv) > 1 and sys.argv[1] == "--drive-summary-csv":
        if len(sys.argv) < 4:
            print("Usage: export_content_batch.py --drive-summary-csv <batch_folder> <url_map.json>")
            sys.exit(1)
        import json

        batch_folder = os.path.abspath(sys.argv[2])
        with open(sys.argv[3], "r", encoding="utf-8") as f:
            url_map = json.load(f)
        md_files = sorted(glob.glob(os.path.join(batch_folder, "week-*.md")))
        content_items = [parse_content_file(fp) for fp in md_files]
        content_items = [i for i in content_items if i["platform"]]
        rows = build_drive_summary_rows(content_items, url_map)
        print(write_drive_summary_csv(rows), end="")
        return

    # Determine batch folder
    if len(sys.argv) > 1:
        batch_folder = sys.argv[1]
    else:
        # Auto-detect from workspace root
        workspace = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        batch_folder = find_latest_batch_folder(workspace)
        if not batch_folder:
            print("Error: No content-batch-* folder found. Provide a path as argument.")
            sys.exit(1)

    batch_folder = os.path.abspath(batch_folder)
    if not os.path.isdir(batch_folder):
        print(f"Error: Directory not found: {batch_folder}")
        sys.exit(1)

    print(f"Processing batch folder: {batch_folder}")

    # Find all markdown files
    md_files = sorted(glob.glob(os.path.join(batch_folder, "*.md")))
    if not md_files:
        print("Error: No .md files found in the batch folder.")
        sys.exit(1)

    print(f"Found {len(md_files)} content files")

    # Parse all files
    content_items = []
    for filepath in md_files:
        try:
            item = parse_content_file(filepath)
            if item["platform"]:
                content_items.append(item)
                print(f"  Parsed: {item['source_file']} -> {item['platform']}")
            else:
                print(f"  Skipped (no platform detected): {os.path.basename(filepath)}")
        except Exception as e:
            print(f"  Error parsing {os.path.basename(filepath)}: {e}")

    if not content_items:
        print("Error: No valid content files parsed.")
        sys.exit(1)

    # Group summary
    platforms = {}
    for item in content_items:
        platforms.setdefault(item["platform"], 0)
        platforms[item["platform"]] += 1
    print(f"\nContent breakdown:")
    for platform, count in sorted(platforms.items()):
        print(f"  {platform}: {count} pieces")

    # Extract batch date for filename
    batch_date = extract_batch_date(batch_folder)

    # Create workbook
    wb = create_workbook(content_items)

    # Save
    output_filename = f"content-batch-summary-{batch_date}.xlsx"
    output_path = os.path.join(batch_folder, output_filename)
    wb.save(output_path)
    print(f"\nExported to: {output_path}")


if __name__ == "__main__":
    main()
